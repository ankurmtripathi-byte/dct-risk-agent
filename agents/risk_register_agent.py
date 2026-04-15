import io
import json
import sqlite3
from datetime import datetime

import anthropic
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from flask import Blueprint, jsonify, render_template, request, send_file

from config import ANTHROPIC_API_KEY, MODEL_ANALYSIS, MODEL_FAST
from database import get_db

risk_register_bp = Blueprint("risk_register", __name__)

LEVEL_PREFIX = {
    "enterprise": "ENT",
    "affiliate":  "AFF",
    "department": "DEP",
    "event":      "EVT",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _next_risk_id(level: str, conn) -> str:
    prefix = LEVEL_PREFIX.get(level, "RSK")
    row = conn.execute(
        "SELECT COUNT(*) AS c FROM risks WHERE level = ?", (level,)
    ).fetchone()
    return f"{prefix}-{(row['c'] or 0) + 1:03d}"


def _risk_row(row) -> dict:
    d = dict(row)
    try:
        d["event_context"] = json.loads(d.get("event_context") or "{}")
    except (ValueError, TypeError):
        d["event_context"] = {}
    return d


def _strip_fences(text: str) -> str:
    """Remove markdown code fences from a string."""
    t = text.strip()
    if t.startswith("```"):
        lines = t.split("\n")
        t = "\n".join(lines[1:])
        if t.rstrip().endswith("```"):
            t = t[:t.rstrip().rfind("```")]
    return t.strip()


# ── Core Intelligence Functions ───────────────────────────────────────────────

def generate_event_risk_register(event_data: dict) -> dict:
    """
    Two-stage Claude pipeline:
    Stage 1: claude-haiku-4-5 (fast) — analyses event context intelligence.
    Stage 2: claude-opus-4-6 with adaptive thinking — generates 15 risks with
             specific, quantified ai_correlation fields.

    Returns dict with 'risks' list and 'intelligence' context dict.
    """
    api_key = (ANTHROPIC_API_KEY or "").strip()
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY not configured")

    client = anthropic.Anthropic(api_key=api_key)

    event_name    = event_data.get("event_name", "")
    venue         = event_data.get("venue", "")
    event_date    = event_data.get("event_date", "")
    attendance    = event_data.get("attendance", "")
    event_type    = event_data.get("event_type", "")
    has_vip       = event_data.get("has_vip", False)
    international = event_data.get("international", False)
    outdoor       = event_data.get("outdoor", "")
    contractors   = event_data.get("contractors", "")
    notes         = event_data.get("notes", "")

    # ── STAGE 1: Context Intelligence (Haiku — fast) ──────────────────────────
    stage1_prompt = f"""You are a DCT (Department of Culture and Tourism, Abu Dhabi) event risk intelligence analyst.

Analyse this event and return concise risk intelligence as a JSON object.
Return ONLY valid JSON with no markdown fences or preamble.

Event: {event_name}
Venue: {venue}
Date: {event_date}
Attendance: {attendance}
Type: {event_type}
International Guests: {international}
VIP / State Protocol: {has_vip}
Outdoor Elements: {outdoor}
Contractors: {contractors or "Standard vendors"}
Notes: {notes or "None"}

Return this exact structure:
{{
  "venue_risk_factors": "<specific physical, access, and capacity risks for this venue>",
  "seasonal_context": "<UAE weather, humidity, dust storms, Ramadan/Eid proximity, National Day>",
  "vip_protocol_requirements": "<ADNOC protocol, police escort, press exclusion zones, UAE VIP security standards>",
  "crowd_dynamics": "<demographics, cultural sensitivities, F&B rules, queuing behaviour>",
  "regulatory_landscape": "<DCT permit conditions, ADCD safety codes, TDIC rules, noise ordinance>",
  "cross_risk_correlations": "<specific domino sequences e.g. heat stroke -> ambulance delay -> bad press>",
  "historical_precedents": "<relevant incidents at similar Abu Dhabi events in last 5 years>",
  "contractor_risks": "<F&B vendors, AV suppliers, security contractors, temporary structures>",
  "media_reputational_context": "<international media presence, social media amplification, government sensitivity>",
  "mitigation_benchmark": "<ISO 31000 best practices for this event category>"
}}"""

    stage1_msg = client.messages.create(
        model=MODEL_FAST,
        max_tokens=2048,
        messages=[{"role": "user", "content": stage1_prompt}]
    )

    raw1 = stage1_msg.content[0].text.strip()
    try:
        intelligence = json.loads(_strip_fences(raw1))
    except json.JSONDecodeError:
        intelligence = {"context_summary": raw1}

    # ── STAGE 2: Risk Generation (Opus + Adaptive Thinking) ──────────────────
    stage2_prompt = f"""You are the Risk Director for the Department of Culture and Tourism (DCT), Abu Dhabi.

━━━ EVENT BRIEF ━━━
Event:        {event_name}
Venue:        {venue}
Date:         {event_date}
Attendance:   {attendance} guests
Type:         {event_type}
VIP/State:    {has_vip}
International:{international}
Outdoor:      {outdoor}
Contractors:  {contractors or "Standard vendors"}

━━━ INTELLIGENCE CONTEXT ━━━
Venue Risks:      {intelligence.get("venue_risk_factors", "")}
Seasonal:         {intelligence.get("seasonal_context", "")}
VIP Protocol:     {intelligence.get("vip_protocol_requirements", "")}
Crowd Dynamics:   {intelligence.get("crowd_dynamics", "")}
Regulatory:       {intelligence.get("regulatory_landscape", "")}
Risk Cascades:    {intelligence.get("cross_risk_correlations", "")}
Precedents:       {intelligence.get("historical_precedents", "")}
Contractor Risks: {intelligence.get("contractor_risks", "")}
Media Context:    {intelligence.get("media_reputational_context", "")}

━━━ TASK ━━━
Generate exactly 15 risks covering ALL 7 categories (at least 2 per category):
Safety | Security | Operational | Reputational | Compliance | Financial | Environmental

━━━ CRITICAL REQUIREMENT: ai_correlation field ━━━
Each ai_correlation MUST:
1. Name at least one OTHER specific risk category from this register
2. Describe a CONCRETE cascade mechanism — not vague "may impact" language
3. Include QUANTIFIED escalation where possible (e.g. "likelihood jumps 3->5")
4. Reference specific DCT/Abu Dhabi context (venue name, regulation, authority)
5. NEVER write generic text like "this risk may affect other areas"

GOOD example of ai_correlation:
"A crowd surge at {venue}'s coastal terrace (Safety, score 16) triggers ADCD mandatory
 incident report within 2 hours, elevating Compliance breach likelihood 2->4. Simultaneous
 Al Jazeera/BBC coverage pushes Reputational damage impact 3->5 within 4 hours, requiring
 DCT Communications Director immediate activation — cascading total portfolio risk score
 by an estimated 35% and forcing DCT Events Director escalation."

BAD example (forbidden): "This safety risk may impact other risk areas and requires attention."

Return ONLY valid JSON — no markdown fences, no preamble:
{{
  "risks": [
    {{
      "category": "Safety|Security|Operational|Reputational|Compliance|Financial|Environmental",
      "title": "5-8 word specific risk title",
      "description": "2-3 sentences: specific risk and consequences for this event",
      "ai_correlation": "Specific cascade insight with mechanism and quantified escalation — never generic",
      "likelihood": 3,
      "impact": 3,
      "velocity": "Immediate|Short-term|Long-term",
      "mitigation": "2-3 sentences of actionable prevention specific to this event",
      "contingency": "1-2 sentences: response if risk materialises",
      "owner": "Specific DCT or event role title",
      "reviewer": "Oversight role title",
      "kri": "Measurable Key Risk Indicator",
      "kri_threshold": "Specific numeric/observable threshold"
    }}
  ]
}}"""

    # Non-streaming call. Use Sonnet for speed + quality.
    # Fallback to Haiku if Sonnet times out.
    try:
        msg = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=6000,
            messages=[{"role": "user", "content": stage2_prompt}],
            timeout=50.0,
        )
    except Exception:
        msg = client.messages.create(
            model=MODEL_FAST,
            max_tokens=3000,
            messages=[{"role": "user", "content": stage2_prompt}],
            timeout=30.0,
        )

    raw2 = ""
    for block in msg.content:
        if block.type == "text":
            raw2 = block.text.strip()
            break

    result = json.loads(_strip_fences(raw2))
    return {"risks": result.get("risks", []), "intelligence": intelligence}


def cascade_to_hierarchy(event_risks: list, entity_name: str, conn) -> list:
    """
    Escalate high-scoring event risks up the 4-level hierarchy:
      score >= 12 -> department
      score >= 16 -> affiliate
      score >= 20 -> enterprise

    Only promotes each risk to the single highest applicable level.
    Saves new risks to DB with parent_risk_id set and returns them.
    """
    THRESHOLDS = [(20, "enterprise"), (16, "affiliate"), (12, "department")]
    cascaded = []
    now = datetime.utcnow().isoformat()

    for risk in event_risks:
        score          = int(risk.get("risk_score", 0))
        parent_risk_id = risk.get("risk_id", "")

        for threshold, level in THRESHOLDS:
            if score >= threshold:
                new_rid = _next_risk_id(level, conn)
                level_entity = {
                    "enterprise": "DCT Enterprise",
                    "affiliate":  entity_name or "DCT Affiliate",
                    "department": f"{entity_name} Department" if entity_name else "DCT Events Department",
                }.get(level, entity_name)

                ai_corr = (
                    f"Cascaded from event risk {parent_risk_id} (score {score} >= {threshold} threshold). "
                    f"{risk.get('ai_correlation', '')} "
                    f"Elevated to {level} level — requires portfolio-wide monitoring and "
                    f"DCT Risk Director reporting."
                )

                conn.execute("""
                    INSERT INTO risks
                      (risk_id,level,entity_name,category,title,description,ai_correlation,
                       likelihood,impact,risk_score,velocity,mitigation,contingency,
                       owner,reviewer,kri,kri_threshold,status,source,parent_risk_id,
                       created_date,updated_date,event_context)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    new_rid, level, level_entity,
                    risk.get("category", "Operational"),
                    f"[{level[:3].upper()}] {risk.get('title', '')}",
                    risk.get("description", ""),
                    ai_corr,
                    int(risk.get("likelihood", 3)),
                    int(risk.get("impact", 3)),
                    score,
                    risk.get("velocity", "Short-term"),
                    risk.get("mitigation", ""),
                    risk.get("contingency", ""),
                    risk.get("owner", ""),
                    risk.get("reviewer", ""),
                    risk.get("kri", ""),
                    risk.get("kri_threshold", ""),
                    "Open", "Cascaded", parent_risk_id,
                    now, now,
                    json.dumps(
                        risk.get("event_context")
                        if isinstance(risk.get("event_context"), dict) else {}
                    )
                ))
                new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                cascaded.append(_risk_row(
                    conn.execute("SELECT * FROM risks WHERE id=?", (new_id,)).fetchone()
                ))
                break  # only promote to the single highest level

    conn.commit()
    return cascaded


def correlate_with_existing(new_risks: list, conn) -> list:
    """
    Find category and keyword overlaps between newly generated risks
    and existing open risks in the DB.

    Returns a list of correlation finding dicts.
    """
    existing = [dict(r) for r in conn.execute(
        "SELECT id, risk_id, title, category, risk_score FROM risks WHERE status != 'Closed'"
    ).fetchall()]
    if not existing:
        return []

    stop = {
        "a","an","the","of","in","and","or","to","for","at","by",
        "with","from","due","on","is","are","this","that","which"
    }
    correlations = []

    for i, nr in enumerate(new_risks):
        nc  = (nr.get("category") or "").lower()
        nt  = (nr.get("title")    or "").lower()
        ntw = set(nt.split()) - stop

        for ex in existing:
            ec  = (ex.get("category") or "").lower()
            et  = (ex.get("title")    or "").lower()
            etw = set(et.split()) - stop
            shared = ntw & etw

            if nc == ec and shared:
                corr_type = "duplicate_candidate"
                detail    = f"Same category '{nr.get('category')}' + shared keywords: {', '.join(shared)}"
            elif nc == ec:
                corr_type = "same_category"
                detail    = f"Existing open {nr.get('category')} risk in register"
            elif shared:
                corr_type = "keyword_overlap"
                detail    = f"Shared topic keywords: {', '.join(shared)}"
            else:
                continue

            correlations.append({
                "new_risk_index":   i,
                "new_risk_title":   nr.get("title", ""),
                "existing_risk_id": ex["risk_id"],
                "existing_title":   ex["title"],
                "existing_score":   ex["risk_score"],
                "correlation_type": corr_type,
                "detail":           detail,
            })

    return correlations


# ── Routes ────────────────────────────────────────────────────────────────────

@risk_register_bp.route("/risk-register")
def risk_register_page():
    return render_template("risk_register.html")


@risk_register_bp.route("/api/risks/stats")
def risk_stats():
    conn   = get_db()
    total  = conn.execute("SELECT COUNT(*) AS c FROM risks").fetchone()["c"]
    high   = conn.execute("SELECT COUNT(*) AS c FROM risks WHERE risk_score >= 12").fetchone()["c"]
    medium = conn.execute(
        "SELECT COUNT(*) AS c FROM risks WHERE risk_score BETWEEN 6 AND 11"
    ).fetchone()["c"]
    low    = conn.execute("SELECT COUNT(*) AS c FROM risks WHERE risk_score <= 5").fetchone()["c"]
    by_level  = [dict(r) for r in conn.execute(
        "SELECT level, COUNT(*) AS c FROM risks GROUP BY level"
    ).fetchall()]
    by_status = [dict(r) for r in conn.execute(
        "SELECT status, COUNT(*) AS c FROM risks GROUP BY status"
    ).fetchall()]
    conn.close()
    return jsonify(total=total, high=high, medium=medium, low=low,
                   by_level=by_level, by_status=by_status)


@risk_register_bp.route("/api/risks", methods=["GET"])
def list_risks():
    level  = request.args.get("level", "")
    entity = request.args.get("entity", "")
    status = request.args.get("status", "")
    cat    = request.args.get("category", "")

    conn = get_db()
    q, p = "SELECT * FROM risks WHERE 1=1", []
    if level:  q += " AND level = ?";          p.append(level)
    if entity: q += " AND entity_name LIKE ?"; p.append(f"%{entity}%")
    if status: q += " AND status = ?";         p.append(status)
    if cat:    q += " AND category = ?";       p.append(cat)
    q += " ORDER BY risk_score DESC, created_date DESC"
    risks = [_risk_row(r) for r in conn.execute(q, p).fetchall()]
    conn.close()
    return jsonify(risks)


@risk_register_bp.route("/api/risks", methods=["POST"])
def create_risk():
    data  = request.get_json(force=True)
    conn  = get_db()
    now   = datetime.utcnow().isoformat()
    level = data.get("level", "event")
    rid   = _next_risk_id(level, conn)
    l, i  = int(data.get("likelihood", 1)), int(data.get("impact", 1))

    conn.execute("""
        INSERT INTO risks
          (risk_id,level,entity_name,category,title,description,ai_correlation,
           likelihood,impact,risk_score,velocity,mitigation,contingency,
           owner,reviewer,kri,kri_threshold,status,source,parent_risk_id,
           created_date,updated_date,event_context)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        rid, level,
        data.get("entity_name",""), data.get("category",""),
        data.get("title",""), data.get("description",""),
        data.get("ai_correlation",""),
        l, i, l * i,
        data.get("velocity","Short-term"),
        data.get("mitigation",""), data.get("contingency",""),
        data.get("owner",""), data.get("reviewer",""),
        data.get("kri",""), data.get("kri_threshold",""),
        data.get("status","Open"), data.get("source","Manual"),
        data.get("parent_risk_id",""),
        now, now,
        json.dumps(data.get("event_context", {}))
    ))
    conn.commit()
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    risk   = _risk_row(conn.execute("SELECT * FROM risks WHERE id=?", (new_id,)).fetchone())
    conn.close()
    return jsonify(risk), 201


@risk_register_bp.route("/api/risks/<int:rid>", methods=["PUT"])
def update_risk(rid):
    data = request.get_json(force=True)
    conn = get_db()
    now  = datetime.utcnow().isoformat()

    fields = [
        "entity_name","category","title","description","ai_correlation",
        "likelihood","impact","velocity","mitigation","contingency",
        "owner","reviewer","kri","kri_threshold","status","parent_risk_id"
    ]
    sets, params = [], []
    for f in fields:
        if f in data:
            sets.append(f"{f} = ?")
            params.append(data[f])

    row = dict(conn.execute("SELECT likelihood,impact FROM risks WHERE id=?", (rid,)).fetchone())
    l = int(data.get("likelihood", row["likelihood"]))
    i = int(data.get("impact",     row["impact"]))
    sets.append("risk_score = ?");   params.append(l * i)
    sets.append("updated_date = ?"); params.append(now)
    params.append(rid)

    conn.execute(f"UPDATE risks SET {', '.join(sets)} WHERE id=?", params)
    conn.commit()
    risk = _risk_row(conn.execute("SELECT * FROM risks WHERE id=?", (rid,)).fetchone())
    conn.close()
    return jsonify(risk)


@risk_register_bp.route("/api/risks/<int:rid>", methods=["DELETE"])
def delete_risk(rid):
    conn = get_db()
    conn.execute("DELETE FROM risks WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return jsonify(deleted=True)


@risk_register_bp.route("/api/risks/generate", methods=["POST"])
def generate_risks():
    data = request.get_json(force=True)
    if not ANTHROPIC_API_KEY:
        return jsonify(error="ANTHROPIC_API_KEY not configured"), 500

    # Accept both old (entity_name/context) and new (event_name/venue/…) formats
    event_data = {
        "event_name":    data.get("event_name") or data.get("entity_name", ""),
        "venue":         data.get("venue", ""),
        "event_date":    data.get("event_date") or data.get("date", ""),
        "attendance":    str(data.get("attendance", "")),
        "event_type":    data.get("event_type", ""),
        "has_vip":       data.get("has_vip", False),
        "international": data.get("international", False),
        "outdoor":       data.get("outdoor", ""),
        "contractors":   data.get("contractors", ""),
        "notes":         data.get("notes", ""),
    }
    entity_name = event_data["event_name"]

    try:
        result       = generate_event_risk_register(event_data)
        risks_raw    = result["risks"]
        intelligence = result["intelligence"]
    except ValueError as e:
        return jsonify(error=str(e)), 500
    except json.JSONDecodeError as e:
        return jsonify(error=f"AI JSON parse error: {e}"), 500
    except anthropic.AuthenticationError:
        return jsonify(error="Invalid Anthropic API key"), 401
    except anthropic.APIError as e:
        return jsonify(error=f"Anthropic API error: {e}"), 500
    except Exception as e:  # noqa: BLE001
        return jsonify(error=f"Unexpected error: {e}"), 500

    conn     = get_db()
    now      = datetime.utcnow().isoformat()
    ctx_json = json.dumps({**event_data, "intelligence": intelligence})
    saved    = []

    for r in risks_raw:
        rid_str = _next_risk_id("event", conn)
        l, i    = int(r.get("likelihood", 3)), int(r.get("impact", 3))
        conn.execute("""
            INSERT INTO risks
              (risk_id,level,entity_name,category,title,description,ai_correlation,
               likelihood,impact,risk_score,velocity,mitigation,contingency,
               owner,reviewer,kri,kri_threshold,status,source,
               created_date,updated_date,event_context)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            rid_str, "event", entity_name,
            r.get("category",""), r.get("title",""), r.get("description",""),
            r.get("ai_correlation",""),
            l, i, l * i, r.get("velocity","Short-term"),
            r.get("mitigation",""), r.get("contingency",""),
            r.get("owner",""), r.get("reviewer",""),
            r.get("kri",""), r.get("kri_threshold",""),
            "Open", "AI-Generated",
            now, now, ctx_json
        ))
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        saved.append(_risk_row(
            conn.execute("SELECT * FROM risks WHERE id=?", (new_id,)).fetchone()
        ))

    conn.commit()
    correlations = correlate_with_existing(risks_raw, conn)
    conn.close()

    return jsonify(
        risks=saved, count=len(saved),
        intelligence=intelligence,
        correlations=correlations
    )


@risk_register_bp.route("/api/risks/cascade", methods=["POST"])
def cascade_risks():
    data        = request.get_json(force=True)
    risk_ids    = data.get("risk_ids", [])
    entity_name = data.get("entity_name", "")

    if not risk_ids:
        return jsonify(error="No risk_ids provided"), 400

    conn = get_db()
    risks_to_cascade = []
    for db_id in risk_ids:
        row = conn.execute("SELECT * FROM risks WHERE id=?", (db_id,)).fetchone()
        if row:
            risks_to_cascade.append(_risk_row(row))

    if not risks_to_cascade:
        conn.close()
        return jsonify(error="No matching risks found"), 404

    cascaded = cascade_to_hierarchy(risks_to_cascade, entity_name, conn)
    conn.close()
    return jsonify(cascaded=cascaded, count=len(cascaded))


@risk_register_bp.route("/api/risks/export", methods=["GET"])
def export_risks():
    level  = request.args.get("level", "")
    status = request.args.get("status", "")

    conn = get_db()
    q, p = "SELECT * FROM risks WHERE 1=1", []
    if level:  q += " AND level = ?";  p.append(level)
    if status: q += " AND status = ?"; p.append(status)
    q += " ORDER BY level, risk_score DESC"
    risks = [_risk_row(r) for r in conn.execute(q, p).fetchall()]
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Register"

    hdr_fill  = PatternFill("solid", fgColor="0A1628")
    crit_fill = PatternFill("solid", fgColor="FECACA")   # light red
    high_fill = PatternFill("solid", fgColor="FEE2E2")
    med_fill  = PatternFill("solid", fgColor="FFFBEB")
    low_fill  = PatternFill("solid", fgColor="F0FDF4")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    bold_font = Font(bold=True, size=10)

    headers = [
        "Risk ID", "Level", "Entity", "Category", "Title", "Description",
        "AI Correlation", "L", "I", "Score", "Velocity",
        "Mitigation", "Contingency", "Owner", "Reviewer",
        "KRI", "KRI Threshold", "Status", "Source", "Created"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for ri, r in enumerate(risks, 2):
        score = r.get("risk_score", 0)
        fill  = (crit_fill if score >= 20 else
                 high_fill if score >= 12 else
                 med_fill  if score >= 6  else
                 low_fill)

        created = (r.get("created_date") or "")[:10]
        row_data = [
            r.get("risk_id",""), r.get("level",""), r.get("entity_name",""),
            r.get("category",""), r.get("title",""), r.get("description",""),
            r.get("ai_correlation",""),
            r.get("likelihood",""), r.get("impact",""), score,
            r.get("velocity",""),
            r.get("mitigation",""), r.get("contingency",""),
            r.get("owner",""), r.get("reviewer",""),
            r.get("kri",""), r.get("kri_threshold",""),
            r.get("status",""), r.get("source",""), created,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=ri, column=10).font = bold_font

    col_widths = [10, 10, 18, 14, 28, 40, 50, 4, 4, 6, 10, 40, 32, 24, 24, 30, 18, 10, 14, 12]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"DCT_Risk_Register_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )
